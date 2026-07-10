#!/usr/bin/env python3
"""
build_data.py — Refactor the BOM "FWIN ESS Template" workbook into clean,
version-controllable JSON that drives both the browser tool and the ESS skill.

Why this exists
---------------
The source of truth for station metadata is a ~4 MB Excel workbook
("FWIN_ESS_Template_Oct_2025.xlsx"). The ESS proforma sheet uses VLOOKUP over a
"Station Data Sheet" tab to auto-populate station number / WMO / state / delivery
group / facility name / lat / long from a chosen station name, and pulls a few
state-based reference links (weeds / invasive animals / disease outbreak).

This script pulls the same data out of the workbook into:

  data/stations.json    canonical, de-duplicated station list (one entry per
                        station number, facility types merged) with the
                        state-based reference links attached.
  data/dropdowns.json   the standardized statement lists from the
                        "Auto Fill Information" tab (Permits, Biosecurity,
                        Threatened Habitat/Flora/Fauna, Heritage, Invasive
                        plants/animals, Disease) so the report tool offers the
                        exact wording the current proforma uses.
  data/reference/weeds.json     the alphabetical weed reference list.
  data/reference/diseases.json  the disease/pathogen reference notes.
  data/meta.json        build provenance (source file, row counts, timestamp).

Only the *derived* JSON is committed. The raw workbook lives in build/source/
(git-ignored) so the internal template — with its embedded screenshots and
SharePoint links — is never published to GitHub Pages.

Usage
-----
    pip install -r build/requirements.txt
    python build/build_data.py \
        --source build/source/FWIN_ESS_Template.xlsx \
        --out data

Re-run whenever the workbook is updated. The output is deterministic (sorted),
so a clean `git diff` shows exactly what changed between template versions.
"""
from __future__ import annotations

import argparse
import collections
import datetime as _dt
import hashlib
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")  # openpyxl warns about the x14 data-validation ext

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install -r build/requirements.txt")


# --------------------------------------------------------------------------- #
# State normalisation + state-based reference links
# --------------------------------------------------------------------------- #
# The workbook's "Region" field uses BOM delivery regions. Normalise to the
# standard state/territory codes the state-based search tools are keyed on.
REGION_TO_STATE = {
    "NSW": "NSW",
    "QLD": "QLD",
    "VIC": "VIC",
    "SA": "SA",
    "WA": "WA",
    "NT": "NT",
    "ACT": "ACT",
    "TAS": "TAS",
    "TAS/ANT": "TAS",  # Tasmania / Antarctica delivery region
    "HO": "",          # Head Office — resolve from coordinates instead
}

# Canonical, public reference links per state. Seeded from the workbook's
# "Station Data Sheet" (Weeds Reference column) and completed for the states the
# workbook did not cover, so every jurisdiction resolves to a real URL. Edit here
# when a link changes — it flows through to every station on the next build.
STATE_WEEDS_LINK = {
    "QLD": "https://www.business.qld.gov.au/industries/farms-fishing-forestry/agriculture/biosecurity/plants/invasive/restricted",
    "NSW": "https://weeds.org.au/regions/nsw/",
    "VIC": "https://agriculture.vic.gov.au/biosecurity/weeds/weeds-information",
    "SA": "https://www.landscape.sa.gov.au/hf/landscapes-hills-and-fleurieu-stewardship-program/understand-your-responsibilities-as-a-land-manager/pest-plants-and-animals-2/pest-plants",
    "WA": "https://www.agric.wa.gov.au/pests-weeds-diseases/weeds",
    "TAS": "https://nre.tas.gov.au/invasive-species/weeds-index",
    "NT": "https://nt.gov.au/environment/weeds",
    "ACT": "https://www.environment.act.gov.au/parks-conservation/plants-and-animals/pest-plants-and-animals",
}
# Invasive animals + disease outbreak links are national in the workbook.
NATIONAL_INVASIVE_ANIMALS_LINK = "https://www.dcceew.gov.au/environment/invasive-species"
NATIONAL_OUTBREAK_LINK = "https://www.outbreak.gov.au/"

# Rough state bounding boxes to resolve "HO" / coordinate-only sites to a state.
# Heuristic only (boxes overlap near borders); the browser tool lets the user
# override. Ordered most-specific first so ACT wins over NSW.
STATE_BBOXES = [
    ("ACT", -35.92, -35.12, 148.76, 149.40),
    ("TAS", -43.75, -39.10, 143.80, 148.55),
    ("VIC", -39.20, -33.98, 140.96, 150.05),
    ("NSW", -37.51, -28.16, 140.99, 153.64),
    ("QLD", -29.18, -9.90, 137.99, 153.55),
    ("SA", -38.10, -25.99, 128.99, 141.02),
    ("NT", -26.01, -10.90, 128.99, 138.02),
    ("WA", -35.20, -13.68, 112.90, 129.02),
]


def state_from_coords(lat, lon):
    """Best-effort state code from a coordinate. Returns '' if unknown."""
    if lat is None or lon is None:
        return ""
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return ""
    for code, s, n, w, e in STATE_BBOXES:
        if s <= lat <= n and w <= lon <= e:
            return code
    return ""


def resolve_state(region, lat, lon):
    """Normalise the region field to a state code, falling back to coordinates."""
    code = REGION_TO_STATE.get(str(region).strip(), None)
    if code:
        return code
    # Region is blank, HO, or unrecognised — infer from position.
    return state_from_coords(lat, lon)


# --------------------------------------------------------------------------- #
# Workbook helpers
# --------------------------------------------------------------------------- #
def load(path):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def sheet_rows(ws, header_row=1):
    """Yield dict rows keyed by the header row's column labels."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[header_row - 1]]
    out = []
    for r in rows[header_row:]:
        out.append({header[i]: r[i] for i in range(len(header)) if header[i]})
    return header, out


def clean_num(v):
    """Station/WMO numbers come through as ints, floats or strings — normalise."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_coord(v):
    try:
        return round(float(v), 6)
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- #
# Station extraction
# --------------------------------------------------------------------------- #
# The workbook holds stations across several tabs, and no single tab is a
# superset. In particular the FWN "Station Data Sheet" (which drives the
# proforma's autofill) contains ~315 sites — e.g. WOODGATE ALERT / 539251 — that
# the broad "EAMS DATA SET" export does not. We therefore UNION every
# station-bearing tab, keyed by station number, so any site can be picked by name.
#
# Two tabs have a header row and are parsed by column label; the rest are
# header-less exports with a fixed column order captured here (0-indexed).
HEADERLESS_LAYOUTS = {
    # sheet name: {field: column index}
    "FWN Sites":         {"name": 3, "region": 0, "dg": 1, "num": 2, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Manual Stations":   {"name": 0, "region": 1, "dg": 2, "num": 3, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Rainfall":          {"name": 0, "region": 1, "dg": 2, "num": 3, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Offices":           {"name": 0, "region": 1, "dg": 2, "num": 3, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Remote Anemometer": {"name": 0, "region": 1, "dg": 2, "num": 3, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Balloon Stations":  {"name": 3, "region": 0, "dg": 1, "num": 2, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
    "Misc":              {"name": 3, "region": 0, "dg": 1, "num": 2, "wmo": 4, "facility": 5, "lat": 6, "lon": 7, "oa": 8, "ident": 9},
}


def _norm_rows_from_header_sheet(ws, colmap):
    """EAMS / Station Data Sheet: parse by header label into the common shape."""
    _, rows = sheet_rows(ws)
    for x in rows:
        yield {
            "name": x.get(colmap["name"]),
            "region": x.get(colmap["region"]),
            "dg": x.get(colmap["dg"]),
            "num": x.get(colmap["num"]),
            "wmo": x.get(colmap["wmo"]),
            "facility": x.get(colmap["facility"]),
            "lat": x.get(colmap["lat"]),
            "lon": x.get(colmap["lon"]),
            "oa": x.get(colmap.get("oa")) if colmap.get("oa") else None,
            "ident": x.get(colmap.get("ident")) if colmap.get("ident") else None,
        }


def _norm_rows_from_headerless(ws, layout):
    """Categorised tabs (no header): read by fixed column index."""
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        def g(field):
            i = layout.get(field)
            return row[i] if (i is not None and i < len(row)) else None
        yield {k: g(k) for k in ("name", "region", "dg", "num", "wmo", "facility", "lat", "lon", "oa", "ident")}


def build_stations(wb):
    """One canonical record per station number, merged across every tab."""
    by_key = collections.OrderedDict()
    skipped = 0

    # Order matters: EAMS first (richest metadata), then the FWN autofill list
    # (adds missing FWN sites + is authoritative for that network), then the rest.
    sources = []
    if "EAMS DATA SET" in wb.sheetnames:
        sources.append(("EAMS DATA SET", _norm_rows_from_header_sheet(
            wb["EAMS DATA SET"],
            {"name": "Station Name", "region": "Region", "dg": "Delivery Group",
             "num": "Station Num", "wmo": "WMO", "facility": "Facility Name",
             "lat": "LAT", "lon": "LONG", "oa": "Operating Authority (Station)", "ident": "IDENT"})))
    if "Station Data Sheet" in wb.sheetnames:
        sources.append(("Station Data Sheet", _norm_rows_from_header_sheet(
            wb["Station Data Sheet"],
            {"name": "Station Name", "region": "Region", "dg": "Delivery Group",
             "num": "Station Num", "wmo": "WMO", "facility": "Facility Name",
             "lat": "LAT", "lon": "LONG", "oa": "Operating Authority (Station)", "ident": "IDENT"})))
    for sheet, layout in HEADERLESS_LAYOUTS.items():
        if sheet in wb.sheetnames:
            sources.append((sheet, _norm_rows_from_headerless(wb[sheet], layout)))

    for sheet, rows in sources:
        for x in rows:
            name = x.get("name")
            name = str(name).strip() if name is not None else ""
            num = clean_num(x.get("num"))
            lat = clean_coord(x.get("lat"))
            lon = clean_coord(x.get("lon"))
            if not name or lat is None or lon is None:
                skipped += 1
                continue

            key = num or f"{name.upper()}@{lat:.5f},{lon:.5f}"
            facility = (str(x.get("facility")).strip() if x.get("facility") else "")
            region = x.get("region")

            rec = by_key.get(key)
            if rec is None:
                state = resolve_state(region, lat, lon)
                rec = {
                    "name": name,
                    "station_num": num,
                    "wmo": clean_num(x.get("wmo")),
                    "region": (str(region).strip() if region else ""),
                    "state": state,
                    "delivery_group": (str(x.get("dg")).strip() if x.get("dg") else ""),
                    "facility_types": [],
                    "primary_facility": facility,
                    "lat": lat,
                    "lon": lon,
                    "operating_authority": (str(x.get("oa")).strip() if x.get("oa") else ""),
                    "ident": clean_num(x.get("ident")),
                    "refs": {
                        "invasive_plants": STATE_WEEDS_LINK.get(state, ""),
                        "invasive_animals": NATIONAL_INVASIVE_ANIMALS_LINK,
                        "diseases": NATIONAL_OUTBREAK_LINK,
                    },
                    "sources": [sheet],
                }
                by_key[key] = rec
            else:
                # Fill blanks from later tabs; union facility types + provenance.
                if not rec["wmo"]:
                    rec["wmo"] = clean_num(x.get("wmo"))
                if not rec["delivery_group"] and x.get("dg"):
                    rec["delivery_group"] = str(x.get("dg")).strip()
                if not rec["region"] and region:
                    rec["region"] = str(region).strip()
                    if not rec["state"]:
                        rec["state"] = resolve_state(region, lat, lon)
                        rec["refs"]["invasive_plants"] = STATE_WEEDS_LINK.get(rec["state"], "")
                if not rec["operating_authority"] and x.get("oa"):
                    rec["operating_authority"] = str(x.get("oa")).strip()
                if sheet not in rec["sources"]:
                    rec["sources"].append(sheet)
            if facility and facility not in rec["facility_types"]:
                rec["facility_types"].append(facility)

    stations = sorted(by_key.values(), key=lambda r: (r["name"], r["station_num"]))
    for r in stations:
        r["facility_types"].sort()
    return stations, skipped


# --------------------------------------------------------------------------- #
# Dropdown (standardized statement) extraction
# --------------------------------------------------------------------------- #
def _col_values(ws, col_letter, r1, r2):
    out = []
    for r in range(r1, r2 + 1):
        v = ws[f"{col_letter}{r}"].value
        if v is not None and str(v).strip():
            out.append(str(v).strip())
    return out


def build_dropdowns(wb):
    """
    Mirror the ESS proforma's data-validation lists. Ranges are taken from the
    'Auto Fill Information' tab exactly as the x14 dataValidations reference them
    (see build/README.md for the cell-range map).
    """
    ws = wb["Auto Fill Information"]
    d = {
        "permits":            _col_values(ws, "D", 3, 6),
        "biosecurity":        _col_values(ws, "D", 9, 11),
        "threatened_habitat": _col_values(ws, "D", 22, 24),
        "threatened_flora":   _col_values(ws, "D", 25, 27),
        "threatened_fauna":   _col_values(ws, "D", 28, 30),
        "diseases":           _col_values(ws, "D", 31, 32),
        "indigenous_areas":   _col_values(ws, "D", 36, 37),
        "heritage":           _col_values(ws, "D", 39, 40),
        "invasive_plants":    _col_values(ws, "D", 45, 47),
        "invasive_animals":   _col_values(ws, "D", 48, 50),
    }
    # Biosecurity option -> full declaration text (proforma cell A19 IFS logic).
    bio_detail = {}
    for r in (15, 17, 19):
        label = ws[f"B{r}"].value
        text = ws[f"C{r}"].value
        if label and text:
            bio_detail[str(label).strip()] = str(text).strip()
    d["biosecurity_detail"] = bio_detail
    return d


def build_reference(wb):
    weeds = []
    if "Weeds" in wb.sheetnames:
        ws = wb["Weeds"]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip():
                weeds.append(str(v).strip())
    diseases = []
    if "Diseases" in wb.sheetnames:
        ws = wb["Diseases"]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip():
                diseases.append(str(v).strip())
    return {"weeds": weeds}, {"notes": diseases}


# --------------------------------------------------------------------------- #
def write_json(path, obj, compact=False):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        if compact:
            # Generated web asset: minified to keep the fetch small (gzips well
            # on GitHub Pages). Provenance/counts live in meta.json.
            json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
        else:
            json.dump(obj, f, ensure_ascii=False, indent=1, sort_keys=False)
        f.write("\n")
    return os.path.getsize(path)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", default="build/source/FWIN_ESS_Template.xlsx")
    ap.add_argument("--out", default="data")
    args = ap.parse_args()

    if not os.path.exists(args.source):
        sys.exit(
            f"Source workbook not found: {args.source}\n"
            "Drop the BOM FWIN ESS Template into build/source/ and re-run.\n"
            "(The raw workbook is git-ignored on purpose — only the derived\n"
            " JSON in data/ is committed.)"
        )

    with open(args.source, "rb") as f:
        digest = hashlib.sha256(f.read()).hexdigest()

    wb = load(args.source)
    stations, skipped = build_stations(wb)
    dropdowns = build_dropdowns(wb)
    weeds, diseases = build_reference(wb)

    by_state = collections.Counter(s["state"] or "?" for s in stations)
    by_type = collections.Counter(t for s in stations for t in s["facility_types"])

    meta = {
        "generated_utc": _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_file": os.path.basename(args.source),
        "source_sha256": digest,
        "station_count": len(stations),
        "rows_skipped_no_coords": skipped,
        "stations_by_state": dict(sorted(by_state.items())),
        "stations_by_facility_type": dict(sorted(by_type.items())),
        "builder": "build/build_data.py",
    }

    sizes = {}
    sizes["stations.json"] = write_json(os.path.join(args.out, "stations.json"), stations, compact=True)
    sizes["dropdowns.json"] = write_json(os.path.join(args.out, "dropdowns.json"), dropdowns)
    sizes["reference/weeds.json"] = write_json(os.path.join(args.out, "reference", "weeds.json"), weeds)
    sizes["reference/diseases.json"] = write_json(os.path.join(args.out, "reference", "diseases.json"), diseases)
    sizes["meta.json"] = write_json(os.path.join(args.out, "meta.json"), meta)

    print("ESS data build complete")
    print(f"  stations:            {len(stations)}")
    print(f"  skipped (no coords): {skipped}")
    print(f"  by state:            {dict(sorted(by_state.items()))}")
    print("  files:")
    for name, size in sizes.items():
        print(f"    {name:28s} {size/1024:8.1f} KiB")


if __name__ == "__main__":
    main()
